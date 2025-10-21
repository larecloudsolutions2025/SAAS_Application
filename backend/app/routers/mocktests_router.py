from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from app.database import get_db
from app import models
from datetime import datetime
from pydantic import BaseModel
from typing import Dict
import pandas as pd
from io import BytesIO
from fpdf import FPDF
from fastapi.responses import Response
import matplotlib.pyplot as plt
import numpy as np
import tempfile
import json
from datetime import datetime
import os, json
from fastapi import Request
from app.routers.auth import get_current_user
from dotenv import load_dotenv

load_dotenv()
BASE_URL = os.getenv("BASE_URL", "http://localhost:8000").rstrip("/")


router = APIRouter(prefix="/mocktests", tags=["mocktests"])

# ---------------- Schemas ----------------
class SubmitPayload(BaseModel):
    answers: Dict[str, str]

class SavePayload(BaseModel):
    answers: Dict[str, str]
    time_left: int
    current_question: int

# ---------------- Helpers ----------------
def normalize_cols(df: pd.DataFrame):
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

def get_val(row, names):
    for n in names:
        if n in row and pd.notna(row[n]) and str(row[n]).strip() != "":
            return row[n]
    return ""

def read_excel(path: str):
    """
    Safely reads Excel with openpyxl backend.
    Forces all cells (including numeric, % and ratios) to strings,
    unmerges merged cells, and ensures all option columns exist.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    import pandas as pd

    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Unmerge cells so each cell has its own value
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    # Convert to DataFrame treating everything as text
    data = ws.values
    cols = [str(c).strip().lower().replace(" ", "_") for c in next(data)]
    df = pd.DataFrame(data, columns=cols, dtype=str)
    df = df.fillna("")

    # Guarantee presence of option_a–option_e
    for col in ["option_a", "option_b", "option_c", "option_d", "option_e"]:
        if col not in df.columns:
            df[col] = ""

    return df


def make_full_url(path):
    if not path:
        return ""
    path = path.replace("\\", "/")
    if path.startswith("http"):
        return path
    if path.startswith("static/"):
        return f"http://localhost:8000/{path}"
    filename = os.path.basename(path)
    return f"http://localhost:8000/static/mocktest_images/{filename}"





def read_test_excel(file_path):
    """
    Read test excel and return list of rows as dicts.
    Auto-normalizes option columns (A–E) and fills blanks safely.
    """
    df = pd.read_excel(file_path)
    df.columns = [str(c).strip().lower() for c in df.columns]
    rows = df.fillna("").to_dict(orient="records")
    normalized = []

    for i, r in enumerate(rows, start=1):
        # --- Question ID ---
        qid = ""
        for key in ["question_id", "qid", "id"]:
            if key in r and r[key] != "":
                qid = str(int(r[key])) if isinstance(r[key], (int, float)) else str(r[key])
                break
        if not qid:
            qid = str(i)

        # --- Question text / passage ---
        qtext = r.get("question") or r.get("question_text") or r.get("Question") or ""
        passage_id = r.get("passage_id") or r.get("passageid") or ""
        passage_text = r.get("passage_text") or r.get("passage") or ""

        # --- Normalize and collect options ---
        def get_opt(keys):
            for k in keys:
                if k in r and str(r[k]).strip():
                    return str(r[k]).strip()
            return ""

        a = get_opt(["option_a", "a", "option 1", "opt_a", "optiona"])
        b = get_opt(["option_b", "b", "option 2", "opt_b", "optionb", "option b"])
        c = get_opt(["option_c", "c", "option 3", "opt_c", "optionc", "option c"])
        d = get_opt(["option_d", "d", "option 4", "opt_d", "optiond", "option d"])
        e = get_opt(["option_e", "e", "option 5", "opt_e", "optione", "option e"])

        options = [x for x in [a, b, c, d, e] if x != ""]

        correct = r.get("correct_option") or r.get("answer") or r.get("correct") or ""
        section = r.get("section") or r.get("topic") or "General"
        explanation = r.get("explanation", "")
        # ✅ Ensure image paths are converted to full URLs
        question_image = make_full_url(r.get("question_image", ""))
        explanation_image = make_full_url(r.get("explanation_image", ""))
        passage_image = make_full_url(r.get("passage_image", ""))


        normalized.append({
            "question_id": str(qid),
            "question_text": str(qtext),
            "options": options,
            "correct": str(correct).upper(),
            "section": str(section),
            "passage_id": str(passage_id),
            "passage_text": str(passage_text),
            "explanation": str(explanation),
            "explanation_image": explanation_image,
            "question_image": question_image,
            "passage_image": passage_image,
        })


    return normalized


# ---------------- Routes ----------------
@router.get("/full")
def get_full_tests(db: Session = Depends(get_db)):
    tests = db.query(models.MockTestFile).filter(models.MockTestFile.test_type == "full").all()
    return [
        {
            "id": t.id,
            "name": t.name,
            "file_path": t.file_path,
            "duration_minutes": t.duration_minutes,
            "total_questions": getattr(t, "total_questions", None),
            "file_path": t.file_path,
        } for t in tests
    ]

def safe_str(value):
    """Ensures consistent text for all Excel cells."""
    if value is None:
        return ""
    val = str(value).strip()
    if val.lower() in ["nan", "none"]:
        return ""
    return val

def get_options(row):
    """
    Always returns 5 text options (A–E) correctly.
    Handles numeric, ratio, %, and merged-cell blanks safely.
    """
    opts = []
    for col in ["option_a", "option_b", "option_c", "option_d", "option_e"]:
        val = row.get(col, "")
        val = str(val).strip() if val is not None else ""
        # Handle Excel 'nan', 'None', or hidden blanks
        if val.lower() in ["nan", "none"]:
            val = ""
        opts.append(val)
    while len(opts) < 5:
        opts.append("")
    return opts




@router.get("/{test_id}/resume")
def resume_test(test_id: int, db: Session = Depends(get_db)):
    test = db.query(models.MockTestFile).filter(models.MockTestFile.id == test_id).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")

    df = read_excel(test.file_path)
    rows = df.to_dict(orient="records")
    questions = []

    for r in rows:
        qid = str(get_val(r, ["question_id"]))  # strictly from Excel
        if not qid:
            continue  # skip invalid rows

        # ✅ Fetch passage directly from passage_text column
        passage_text = str(get_val(r, ["passage_text", "passage"])) or ""
        passage_image = str(get_val(r, ["passage_image"])) or ""
        
        #opts = [safe_str(get_val(r, [col])) for col in ["option_a", "option_b", "option_c", "option_d", "option_e"]]


        qobj = {
            "question_id": qid,
            "question": str(get_val(r, ["question", "question_text"])) or "",
            "options": get_options(r),
            "correct": str(get_val(r, ["correct_option", "answer"])) or "",
            "explanation": str(get_val(r, ["explanation"])) or "",
            "passage_text": passage_text,
            "section": str(get_val(r, ["section"])) or "General",
            "question_image": make_full_url(get_val(r, ["question_image"])),
            "explanation_image": make_full_url(get_val(r, ["explanation_image"])),
            "passage_image": make_full_url(get_val(r, ["passage_image"])),

        }

        questions.append(qobj)


    return {
        "attempt_id": f"attempt-{test_id}-{int(datetime.utcnow().timestamp())}",
        "test_name": test.name,
        "duration_minutes": test.duration_minutes,
        "questions": questions,
        "answers_snapshot": {},
        "remaining_time": test.duration_minutes * 60,
        "current_question": 0,
    }

@router.post("/{test_id}/submit/{attempt_id}")
def submit_test(
    test_id: int,
    attempt_id: str,
    payload: SubmitPayload,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    print("🧠 Received test submission for test_id:", test_id)
    print("👤 Current User ID:", current_user.id if current_user else "No user")

    test = db.query(models.MockTestFile).filter(models.MockTestFile.id == test_id).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")

    if not current_user or not current_user.id:
        raise HTTPException(status_code=401, detail="User authentication failed")

    df = read_excel(test.file_path)
    rows = df.to_dict(orient="records")

    correct_map, section_map, exp_map, exp_img_map, qimg_map = {}, {}, {}, {}, {}
    for idx, r in enumerate(rows, start=1):
        qid = str(int(get_val(r, ["question_id"]) or idx))
        correct_map[qid] = str(get_val(r, ["correct_option", "answer"]) or "").upper()
        section_map[qid] = str(get_val(r, ["section"]) or "General")
        exp_map[qid] = str(get_val(r, ["explanation"]) or "")
        exp_img_map[qid] = str(get_val(r, ["explanation_image"]) or "")
        qimg_map[qid] = make_full_url(str(get_val(r, ["question_image"]) or ""))

    total = len(correct_map)
    correct_count = 0
    wrong_count = 0
    answers = payload.answers or {}
    details = []
    section_stats = {}

    # ✅ Evaluate questions with section-wise tracking
    for qid, correct in correct_map.items():
        sel = answers.get(qid, "")
        section = section_map[qid]
        is_correct = sel.upper() == correct if sel else False

        # Initialize section stats
        if section not in section_stats:
            section_stats[section] = {
                "attempted": 0,
                "correct": 0,
                "wrong": 0,
                "unattempted": 0,
                "marks": 0.0,
            }

        if not sel:
            section_stats[section]["unattempted"] += 1
        else:
            section_stats[section]["attempted"] += 1
            if is_correct:
                correct_count += 1
                section_stats[section]["correct"] += 1
                section_stats[section]["marks"] += 1  # +1 mark for correct
            else:
                wrong_count += 1
                section_stats[section]["wrong"] += 1
                section_stats[section]["marks"] -= 0.25  # -0.25 mark for wrong

        details.append({
            "question_id": qid,
            "selected": sel,
            "correct": correct,
            "is_correct": is_correct,
            "explanation": exp_map[qid],
            "explanation_image": exp_img_map[qid],
            "question_image": qimg_map[qid],
            "section": section,
        })

        print(f"🖼️  Question {qid} image: {get_val(r, ['question_image'])}")


    # ✅ Apply global score calculation (based on total correct/wrong)
    score = round(correct_count * 1 - (wrong_count * 0.25), 2)
    percentage = round((score / total) * 100, 2) if total > 0 else 0.0

    # ✅ Store result
    try:
        result = models.UserResult(
            user_id=current_user.id,
            mocktest_id=test_id,
            score=score,
            total_questions=total,
            percentage=str(percentage),
            details=json.dumps({
                "questions": details,
                "sections": section_stats
            }),
            status="completed",
            submitted_at=datetime.utcnow(),
        )
        db.add(result)
        db.commit()
        db.refresh(result)
        print(f"✅ Inserted result successfully: {result.id}")
        print(f"📊 Correct: {correct_count}, Wrong: {wrong_count}, Total: {total}, Score: {score}")
        return {"msg": "submitted", "result_id": result.id}

    except Exception as e:
        db.rollback()
        import traceback
        print("❌ Error saving user result:", e)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Database insert failed: {str(e)}")

# ✅ FIXED: Correct Summary Endpoint
@router.get("/results/summary")
def summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    results = (
        db.query(models.UserResult)
        .filter(models.UserResult.user_id == current_user.id)
        .order_by(models.UserResult.submitted_at.desc())
        .all()
    )

    summary_data = []
    for r in results:
        summary_data.append({
            "id": r.id,
            "mocktest_id": r.mocktest_id,
            "score": r.score,
            "total_questions": r.total_questions,
            "percentage": r.percentage,
            "status": r.status,
            "submitted_at": r.submitted_at,
        })

    return summary_data

@router.get("/result/{test_id}")
def result(
    test_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    res = (
        db.query(models.UserResult)
        .filter(
            models.UserResult.mocktest_id == test_id,
            models.UserResult.user_id == current_user.id
        )
        .order_by(models.UserResult.id.desc())
        .first()
    )
    if not res:
        raise HTTPException(status_code=404, detail="No result found for this user")
    return {
        "mocktest_id": res.mocktest_id,
        "score": res.score,
        "total_questions": res.total_questions,
        "percentage": res.percentage,
        "submitted_at": res.submitted_at,
        "details": (
            res.details
            if isinstance(res.details, (dict, list))
            else json.loads(res.details)
        ),
    }

@router.get("/result/{result_id}/preview")
def preview_result(
    result_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    result = db.query(models.UserResult).filter(models.UserResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")

    # Parse JSON safely
    try:
        parsed = json.loads(result.details) if isinstance(result.details, str) else result.details
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse details: {e}")

    if isinstance(parsed, dict):
        stored_questions = parsed.get("questions", [])
    elif isinstance(parsed, list):
        stored_questions = parsed
    else:
        stored_questions = []

    # Read master Excel
    mocktest = db.query(models.MockTestFile).filter(models.MockTestFile.id == result.mocktest_id).first()
    if not mocktest:
        raise HTTPException(status_code=404, detail="Mocktest file not found")

    if not os.path.exists(mocktest.file_path):
        raise HTTPException(status_code=500, detail="Mocktest file missing")

    excel_rows = read_test_excel(mocktest.file_path)
    excel_map = {str(r.get("question_id")): r for r in excel_rows}

    # Merge stored details with Excel and rebuild options if missing
    merged = []
    for q in stored_questions:
        qid = str(q.get("question_id", ""))
        base = excel_map.get(qid, {})

        # 🧩 Fallback: construct options if Excel has A/B/C/D/E columns
        options = base.get("options")
        if not options or not isinstance(options, list) or len(options) == 0:
            options = [
                base.get("option_a") or base.get("A") or base.get("Option A") or "",
                base.get("option_b") or base.get("B") or base.get("Option B") or "",
                base.get("option_c") or base.get("C") or base.get("Option C") or "",
                base.get("option_d") or base.get("D") or base.get("Option D") or "",
                base.get("option_e") or base.get("E") or base.get("Option E") or "",
            ]
            # remove empty options if Excel used fewer than 5
            options = [opt for opt in options if opt.strip()]

        merged.append({
            "question_id": qid,
            "question_text": base.get("question_text", ""),
            "options": options,
            "selected": q.get("selected", ""),
            "correct": base.get("correct") or base.get("correct_option") or q.get("correct", ""),
            "is_correct": q.get("is_correct", False),
            "section": base.get("section", q.get("section", "General")),
            "passage_id": base.get("passage_id", ""),
            "passage_text": base.get("passage_text", ""),
            "explanation": base.get("explanation", q.get("explanation", "")),
            "question_image": q.get("question_image") or base.get("question_image", ""),
            "explanation_image": q.get("explanation_image") or base.get("explanation_image", "")
        })

    # ✅ Compute section-wise summary dynamically
    section_map = {}
    for q in merged:
        sec = q.get("section", "General")
        sel = q.get("selected", "")
        corr = q.get("correct", "")
        is_corr = q.get("is_correct", False)
        if sec not in section_map:
            section_map[sec] = {"attempted": 0, "correct": 0, "wrong": 0, "unattempted": 0}
        if not sel:
            section_map[sec]["unattempted"] += 1
        elif is_corr:
            section_map[sec]["attempted"] += 1
            section_map[sec]["correct"] += 1
        else:
            section_map[sec]["attempted"] += 1
            section_map[sec]["wrong"] += 1

    sections_summary = []
    for sec, data in section_map.items():
        total = data["attempted"] + data["unattempted"]
        data["marks"] = round(data["correct"] * 1 - data["wrong"] * 0.25, 2)
        data["percentage"] = round((data["correct"] / total) * 100, 2) if total else 0
        sections_summary.append({
            "section_name": sec,
            **data
        })

    # ---------------- Topper vs You Stats ----------------
    all_results = (
        db.query(models.UserResult)
        .filter(models.UserResult.mocktest_id == result.mocktest_id)
        .order_by(models.UserResult.score.desc())
        .all()
    )

    total_users = len(all_results)
    topper_score = all_results[0].score if total_users > 0 else 0
    your_score = result.score
    rank = next((i + 1 for i, r in enumerate(all_results) if r.id == result.id), total_users)
    percentile = round(((total_users - rank) / total_users) * 100, 2) if total_users > 0 else 0

    performance_band = (
        "Top 10%" if percentile >= 90
        else "Top 25%" if percentile >= 75
        else "Top 50%" if percentile >= 50
        else "Needs Improvement"
    )

    # ✅ Final response
    return {
        "id": result.id,
        "mocktest_id": result.mocktest_id,
        "score": result.score,
        "total_questions": result.total_questions,
        "percentage": result.percentage,
        "submitted_at": result.submitted_at,
        "questions": merged,
        "sections_summary": sections_summary,
        "analytics": {
            "rank": rank,
            "total_users": total_users,
            "percentile": percentile,
            "topper_score": topper_score,
            "performance_band": performance_band
        }
    }



@router.get("/result/{result_id}/download")
def download_result_pdf(
    result_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    import base64

    result = db.query(models.UserResult).filter(models.UserResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")

    mocktest = db.query(models.MockTestFile).filter(models.MockTestFile.id == result.mocktest_id).first()
    mocktest_name = getattr(mocktest, "mocktest_name", None) or getattr(mocktest, "name", f"Mock Test {result.mocktest_id}")

    # ✅ Safely parse result.details
    try:
        parsed = json.loads(result.details) if isinstance(result.details, str) else result.details
        if isinstance(parsed, list):
            questions = parsed
            sections = {"General": {"attempted": 0, "correct": 0, "wrong": 0, "unattempted": 0}}
        elif isinstance(parsed, dict):
            questions = parsed.get("questions", [])
            sections = parsed.get("sections", {"General": {"attempted": 0, "correct": 0, "wrong": 0, "unattempted": 0}})
        else:
            questions, sections = [], {"General": {"attempted": 0, "correct": 0, "wrong": 0, "unattempted": 0}}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse result details: {e}")

    # ✅ Calculate section-wise statistics dynamically
    for q in questions:
        sec = q.get("section", "General")
        if sec not in sections:
            sections[sec] = {"attempted": 0, "correct": 0, "wrong": 0, "unattempted": 0}
        sel = q.get("selected", "")
        correct = q.get("correct", "")
        if not sel:
            sections[sec]["unattempted"] += 1
        elif sel == correct:
            sections[sec]["attempted"] += 1
            sections[sec]["correct"] += 1
        else:
            sections[sec]["attempted"] += 1
            sections[sec]["wrong"] += 1

    # ✅ Calculate marks and totals
    for sec, stats in sections.items():
        stats["marks"] = round(stats["correct"] * 1 - stats["wrong"] * 0.25, 2)
        stats["total_marks"] = stats["attempted"] + stats["unattempted"]

    # ✅ Prepare chart data
    section_names = list(sections.keys())
    correct = [s["correct"] for s in sections.values()]
    wrong = [s["wrong"] for s in sections.values()]
    unattempted = [s["unattempted"] for s in sections.values()]
    marks = [s["marks"] for s in sections.values()]

    # --- Chart 1: Section Performance ---
    plt.figure(figsize=(6, 4))
    plt.bar(section_names, correct, label="Correct", color="#28a745")
    plt.bar(section_names, wrong, bottom=correct, label="Wrong", color="#dc3545")
    plt.bar(section_names, unattempted, bottom=[c + w for c, w in zip(correct, wrong)],
            label="Unattempted", color="#ffc107")
    plt.title("Section-wise Performance (Questions Distribution)")
    plt.xlabel("Sections")
    plt.ylabel("Number of Questions")
    plt.legend()
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format="png")
    buf.seek(0)
    bar_chart_b64 = base64.b64encode(buf.read()).decode("utf-8")
    plt.close()

    # --- Chart 2: Pie Chart ---
    total_correct = sum(correct)
    total_wrong = sum(wrong)
    total_unattempted = sum(unattempted)
    plt.figure(figsize=(4, 4))
    plt.pie(
        [total_correct, total_wrong, total_unattempted],
        labels=["Correct", "Wrong", "Unattempted"],
        autopct="%1.1f%%",
        colors=["#28a745", "#dc3545", "#ffc107"],
        startangle=90,
    )
    plt.title("Overall Performance Distribution")
    buf = BytesIO()
    plt.savefig(buf, format="png")
    buf.seek(0)
    pie_chart_b64 = base64.b64encode(buf.read()).decode("utf-8")
    plt.close()

    # --- Chart 3: Marks per Section (Horizontal) ---
    plt.figure(figsize=(6, 3))
    plt.barh(section_names, marks, color="#007bff")
    plt.title("Marks Scored per Section")
    plt.xlabel("Marks")
    plt.ylabel("Sections")
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format="png")
    buf.seek(0)
    marks_chart_b64 = base64.b64encode(buf.read()).decode("utf-8")
    plt.close()

    # ✅ Performance Rating
    percentage = float(result.percentage)
    if percentage >= 80:
        performance = "Excellent"
    elif percentage >= 60:
        performance = "Good"
    else:
        performance = "Needs Improvement"

    # --- PDF Generation ---
    pdf = FPDF()
    pdf.add_page()

    # Header banner + logo
    pdf.set_fill_color(255, 140, 0)
    pdf.rect(0, 0, 210, 25, "F")

    logo_path = "app/static/LCS_logo.jpg"
    if os.path.exists(logo_path):
        try:
            pdf.image(logo_path, x=10, y=3, w=25)
        except Exception as e:
            print(f"⚠️ Logo load issue: {e}")
    else:
        print("⚠️ Logo not found, skipping logo.")

    pdf.set_xy(40, 5)
    pdf.set_font("Arial", "B", 16)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, "LARE CLOUD SOLUTIONS", ln=True, align="L")

    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 10, "Elevating Ideas. Empowering Futures.", ln=True, align="L")

    # Candidate Info
    pdf.set_text_color(0, 0, 0)
    pdf.ln(20)
    pdf.set_font("Arial", "", 12)
    candidate_name = getattr(current_user, "name", getattr(current_user, "username", "Unknown Candidate"))
    pdf.cell(0, 10, f"Candidate: {candidate_name}", ln=True)
    pdf.cell(0, 10, f"Email: {getattr(current_user, 'email', 'N/A')}", ln=True)
    pdf.cell(0, 10, f"Mock Test Name: {mocktest_name}", ln=True)
    pdf.cell(0, 10, f"Mock Test ID: {result.mocktest_id}", ln=True)
    pdf.cell(0, 10, f"Date: {result.submitted_at.strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(5)

    # --- Section Table ---
    pdf.set_font("Arial", "B", 12)
    pdf.cell(40, 10, "Section", border=1, align="C")
    pdf.cell(25, 10, "Attempted", border=1, align="C")
    pdf.cell(25, 10, "Correct", border=1, align="C")
    pdf.cell(25, 10, "Wrong", border=1, align="C")
    pdf.cell(30, 10, "Unattempted", border=1, align="C")
    pdf.cell(25, 10, "Marks", border=1, align="C")
    pdf.cell(25, 10, "Total Qs", border=1, ln=True, align="C")

    pdf.set_font("Arial", "", 12)
    for section, stats in sections.items():
        pdf.cell(40, 10, section, border=1)
        pdf.cell(25, 10, str(stats["attempted"]), border=1, align="C")
        pdf.cell(25, 10, str(stats["correct"]), border=1, align="C")
        pdf.cell(25, 10, str(stats["wrong"]), border=1, align="C")
        pdf.cell(30, 10, str(stats["unattempted"]), border=1, align="C")
        pdf.cell(25, 10, str(stats["marks"]), border=1, align="C")
        pdf.cell(25, 10, str(stats["total_marks"]), border=1, ln=True, align="C")

    # ✅ Section-wise Total Marks Summary
    pdf.ln(8)
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 10, "Section-wise Total Marks Summary", ln=True)
    pdf.set_font("Arial", "", 12)
    for sec, stats in sections.items():
        pdf.cell(0, 8, f"Section: {sec} - {stats['total_marks']} Questions | Marks Secured: {stats['marks']}", ln=True)

    # ✅ Final Overall Summary
    total_possible_marks = sum([s["total_marks"] for s in sections.values()])
    pdf.ln(10)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Total Score: {result.score} / {result.total_questions}", ln=True)
    pdf.cell(0, 10, f"Percentage: {result.percentage}%", ln=True)
    pdf.cell(0, 10, f"Performance: {performance}", ln=True)
    pdf.cell(0, 10, f"Overall Marks Secured: {result.score} / {total_possible_marks}", ln=True)
    pdf.ln(10)

    # Charts
    pdf.image(BytesIO(base64.b64decode(bar_chart_b64)), x=15, w=180)
    pdf.ln(10)
    pdf.image(BytesIO(base64.b64decode(marks_chart_b64)), x=15, w=180)
    pdf.ln(10)
    pdf.image(BytesIO(base64.b64decode(pie_chart_b64)), x=60, w=80)
    pdf.ln(10)

    # Footer
    pdf.set_font("Arial", "I", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 10, "Generated by Smart MockTest System © 2025 | Powered by Lare Cloud Solutions", ln=True, align="C")

    # ✅ Flush and send PDF response
    pdf_buffer = BytesIO()
    pdf.output(pdf_buffer)
    pdf_data = pdf_buffer.getvalue()
    pdf_buffer.close()

    headers = {"Content-Disposition": f"attachment; filename=Result_Report_{result.id}.pdf"}
    return Response(content=pdf_data, media_type="application/pdf", headers=headers)


@router.get("/subject")
def get_subject_tests(db: Session = Depends(get_db)):
    tests = (
        db.query(models.MockTestFile)
        .filter(models.MockTestFile.test_type == "subject", models.MockTestFile.is_active == True)
        .order_by(models.MockTestFile.subject.asc())
        .all()
    )
    return [
        {
            "id": t.id,
            "name": t.name,
            "subject": t.subject,
            "duration_minutes": t.duration_minutes,
            "total_questions": t.total_questions,
        }
        for t in tests
    ]

